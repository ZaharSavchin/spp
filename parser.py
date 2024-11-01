import time
from selenium.webdriver.support import expected_conditions as EC
from aiogram import Bot
from aiogram.types import FSInputFile
from selenium import webdriver
from selenium.webdriver.common.by import By
import asyncio

from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromiumService
from openpyxl import load_workbook, Workbook
from pass_captcha import tru
from id_token import admin_id, token


bot = Bot(token=token, parse_mode='HTML')


async def main_search(urls_, message):
    workbook = Workbook()
    # workbook = load_workbook(filename='wb1.xlsx')
    # sheet = workbook.active
    # arts = []
    # for cell in sheet['A']:
    #     arts.append(cell.value)

    # options_chrome = webdriver.ChromeOptions()
    # options_chrome.add_extension('guru.crx')

    articles = ['артикул']
    price_with_spp = ['цена']
    price = ['цена с кошельком вб']
    fidbacks = ['отзывы']
    names = ['название карточки']
    colors = ['цвет']
    models = ['модель']
    memorys = ['память']
    oper_memorys = ['оперативная память']
    sellers = ['продавец']
    links = ['ссылка']

    arts = []

    with webdriver.Chrome() as browser:
        for url in urls_:
            group_arts = []
            for ur in url:
                browser.get(ur)
                browser.maximize_window()
                await asyncio.sleep(5)
                hrefs = browser.find_elements(By.TAG_NAME, 'article')
                url_arts = []
                for href in hrefs:
                    url_arts.append(href.get_attribute('data-nm-id'))
                group_arts.append(url_arts)
                await asyncio.sleep(5)
            arts.append(group_arts)

    print(arts)

    # links = [f'https://www.wildberries.ru/catalog/{art}/detail.aspx' for art in arts]

    def open_widget(browser):
        browser.find_element(By.CLASS_NAME, "WidgetButton_widgetButtonOpenIcon__uuM6O").click()

    with webdriver.Chrome() as browser:
        # url = f'https://www.wildberries.ru/catalog/{arts[0]}/detail.aspx'
        #
        # browser.get(url)
        # await asyncio.sleep(2)
        #
        # browser.switch_to.window(browser.window_handles[1])
        #
        # try:
        #     tru(browser)
        # except Exception as err:
        #     print(err)
        #
        # browser.switch_to.window(browser.window_handles[0])
        #
        # browser.get(url)
        # open_widget(browser)

        construction = [[['art', 'art'], ['art', 'art']], [['art', 'art'], ['art', 'art']]]
        counter = 1
        for art_group in arts:
            articles = ['артикул']
            price_with_spp = ['цена']
            price = ['цена с кошельком вб']
            fidbacks = ['отзывы']
            names = ['название карточки']
            colors = ['цвет']
            models = ['модель']
            memorys = ['память']
            oper_memorys = ['оперативная память']
            sellers = ['продавец']
            links = ['ссылка']

            print(art_group)
            sheet = workbook.create_sheet(title=f'list {counter}')
            counter += 1

            for art_link in art_group:
                articles.extend([' ', ' '])
                price_with_spp.extend([' ', ' '])
                price.extend([' ', ' '])
                fidbacks.extend([' ', ' '])
                names.extend([' ', ' '])
                colors.extend([' ', ' '])
                models.extend([' ', ' '])
                memorys.extend([' ', ' '])
                oper_memorys.extend([' ', ' '])
                sellers.extend([' ', ' '])
                links.extend([' ', ' '])

                for art in art_link:
                    if art:

                        articles.append(art)
                        links.append(f'https://www.wildberries.ru/catalog/{art}/detail.aspx')

                        browser.get(f'https://www.wildberries.ru/catalog/{art}/detail.aspx')
                        browser.maximize_window()
                        await asyncio.sleep(3)

                        try:
                            price_wb = browser.find_element(By.XPATH,
                                                            "/html/body/div[1]/main/div[2]/div/div[3]/div/div[3]/div[14]/div/div[1]/div[1]/div[1]/div/div/p/span/ins").text

                            price_wb = ''.join(price_wb.split(' ')[:-1])
                            price_wb = int(price_wb)
                        except Exception as err:
                            price_wb = None

                        try:
                            price_no_spp = browser.find_element(By.XPATH,
                                                                "/html/body/div[1]/main/div[2]/div/div[3]/div/div[3]/div[14]/div/div[1]/div[1]/div/div/div/p/span/span").text
                            price_no_spp = ''.join(price_no_spp.split(' ')[:-1])
                            price_no_spp = int(price_no_spp)

                        except Exception as err:
                            price_no_spp = None


                        try:
                            seller_1 = '/html/body/div[1]/main/div[2]/div/div[3]/div/div[3]/div[14]/div/div[1]/div[7]/section/div/div/div/a[1]'
                            seller_2 = '/html/body/div[1]/main/div[2]/div/div[3]/div/div[3]/div[14]/div/div[1]/div[7]/section/div/div/div/a[1]'
                            seller_3 = '//*[@id="1fb3c78a-44a5-beca-91c3-f7645a5ba4cb"]/div[3]/div[14]/div/div[1]/div[7]/section/div/div/div/a[1]'
                            seller_4 = '/html/body/div[1]/main/div[2]/div/div[3]/div/div[3]/div[14]/div/div[1]/div[7]/section/div[2]/div/div/a'
                            seller_5 = '//*[@id="556e436c-ed4f-8855-2b04-97c70477eab3"]/div[3]/div[14]/div/div[1]/div[7]/section/div[2]/div/div/a'
                            # seller = WebDriverWait(browser, 10).until(EC.visibility_of_element_located((By.CLASS_NAME, 'seller-info__name')))
                            seller = browser.find_element(By.XPATH, "/html/body/div[1]/main/div[2]/div/div[3]/div/div[3]/div[14]/div/div[1]/div[7]/section/div[2]/div/div/a/span").text

                            sellers.append(seller)

                        except Exception as err:
                            seller = None
                            sellers.append(None)

                        try:
                            sold_out = browser.find_element(By.XPATH,
                                                            '/html/body/div[1]/main/div[2]/div/div[3]/div/div[3]/div[2]/div/p/span')
                            price_with_spp.append(' ')
                            price.append(' ')
                        except Exception as err:
                            price_with_spp.append(price_wb)
                            price.append(price_no_spp)

                        # try:
                        #     sold_items = browser.find_element(By.XPATH, "/html/body/div[7]/div[1]/div[2]/div[3]/div[3]/b/span").text
                        #     sold_items = ''.join(sold_items.split(' ')[:-1])
                        #     sold.append(int(sold_items))
                        # except Exception as err:
                        #     sold_items = None
                        #     sold.append(None)

                        try:
                            backs = browser.find_element(By.XPATH,
                                                         "/html/body/div[1]/main/div[2]/div/div[3]/div/div[4]/div[1]/ul/li[1]/button/span").text

                            backs = ''.join(backs.split(' '))
                            fidbacks.append(int(backs))
                        except Exception as err:
                            backs = None
                            fidbacks.append(None)

                        try:
                            name = browser.find_element(By.XPATH,
                                                        '/html/body/div[1]/main/div[2]/div/div[3]/div/div[3]/div[9]/div[1]/h1').text
                            names.append(name)
                        except Exception as err:
                            name = None
                            names.append(None)

                        try:
                            model = browser.find_element(By.XPATH,
                                                         '/html/body/div[1]/main/div[2]/div/div[3]/div/div[3]/div[10]/div[1]/div[2]/table/tbody/tr[2]/td').text
                            models.append(model)
                        except Exception as err:
                            model = None
                            models.append(None)




                        try:
                            button = browser.find_element(By.XPATH,
                                                          '/html/body/div[1]/main/div[2]/div/div[3]/div/div[3]/div[10]/button').click()
                        except Exception as err:
                            print('no button')


                        try:
                            color = browser.find_element(By.CSS_SELECTOR,
                                                         'body > div.popup.popup-product-details.shown > div > div.product-details > div.product-params > table:nth-child(4) > tbody > tr > td').text
                            colors.append(color)
                        except Exception as err:
                            color = None
                            colors.append(None)

                        try:
                            memory = browser.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/div[2]/table[4]/tbody/tr[1]/td').text
                            memorys.append(memory)
                        except Exception as err:
                            memory = None
                            memorys.append(None)

                        try:
                            oper_memory = browser.find_element(By.XPATH,
                                                               '/html/body/div[1]/div/div[1]/div[2]/table[4]/tbody/tr[2]/td').text
                            oper_memorys.append(oper_memory)
                        except Exception as err:
                            oper_memory = None
                            oper_memorys.append(None)

            for i in range(len(articles)):
                sheet.cell(row=i + 2, column=1, value=articles[i])

            for i in range(len(price_with_spp)):
                sheet.cell(row=i + 2, column=2, value=price_with_spp[i])

            for i in range(len(price)):
                sheet.cell(row=i + 2, column=3, value=price[i])

            # for i in range(len(sold)):
            #     sheet.cell(row=i + 2, column=4, value=sold[i])

            for i in range(len(fidbacks)):
                sheet.cell(row=i + 2, column=5, value=fidbacks[i])

            for i in range(len(colors)):
                sheet.cell(row=i + 2, column=6, value=colors[i])

            for i in range(len(memorys)):
                sheet.cell(row=i + 2, column=7, value=memorys[i])

            for i in range(len(oper_memorys)):
                sheet.cell(row=i + 2, column=8, value=oper_memorys[i])

            for i in range(len(models)):
                sheet.cell(row=i + 2, column=9, value=models[i])

            for i in range(len(names)):
                sheet.cell(row=i + 2, column=10, value=names[i])

            for i in range(len(sellers)):
                sheet.cell(row=i + 2, column=11, value=sellers[i])

            for i in range(len(links)):
                sheet.cell(row=i + 2, column=12, value=links[i])



    workbook.save(filename='wb.xlsx')
    file = FSInputFile('wb.xlsx')
    # await message.answer_document(file)
    for admin in admin_id:
        await bot.send_document(admin, file)


async def main_funck(urls_, message):
    print('start')
    for url in urls_:
        print(url)
    counter = 1

    await main_search(urls_, message)
    counter += 1
    await asyncio.sleep(120)




